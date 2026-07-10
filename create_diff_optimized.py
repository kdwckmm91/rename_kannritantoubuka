"""
差分処理スクリプト（管理番号キー結合版）

【機能1】output/<ファイル名>.xlsx の変更セルを直接赤字にする
【機能2】変更行のみを抽出した output/<ファイル名>_変更箇所.xlsx を作成する

source/ と output/ で同名の .xlsx ファイルを自動検出して比較。
「管理番号」列をキーに内部結合し、TARGET_COLUMNS が異なる行を変更行として扱う。

事前に update_ledgers.py を実行して output/ にファイルを生成しておくこと。
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BASE_DIR   = Path(r"C:\Temp\ツール\sisantantou")
SOURCE_DIR = BASE_DIR / "source"
OUTPUT_DIR = BASE_DIR / "output"

KEY_COLUMN      = "管理番号"
CONTEXT_COLUMNS = ["設置場所", "組織管理区分１名"]
TARGET_COLUMNS  = ["利用者名", "管理担当部課", "管理担当部課名"]
READ_COLUMNS    = [KEY_COLUMN] + CONTEXT_COLUMNS + TARGET_COLUMNS


def values_differ(left, right) -> bool:
    """NaN を考慮した値の比較"""
    if pd.isna(left) and pd.isna(right):
        return False
    if pd.isna(left) or pd.isna(right):
        return True
    return left != right


logger.info("=" * 80)
logger.info("差分処理開始（管理番号キー結合版）")
logger.info("=" * 80)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# source/ の .xlsx のうち output/ に同名ファイルがあるペアを自動検出
pairs = []
for src in sorted(SOURCE_DIR.glob("*.xlsx")):
    dst = OUTPUT_DIR / src.name
    if dst.exists():
        diff_out = OUTPUT_DIR / src.name.replace(".xlsx", "_変更箇所.xlsx")
        pairs.append((src, dst, diff_out))

if not pairs:
    logger.warning("比較対象のファイルペアが見つかりません。")
    logger.warning("先に update_ledgers.py を実行して output/ にファイルを生成してください。")
else:
    logger.info(f"{len(pairs)} 件のファイルペアを検出")

for source_file, updated_file, diff_output_file in pairs:
    logger.info("-" * 60)
    logger.info(f"処理中: {source_file.name}")

    read_filter = lambda col: col in READ_COLUMNS
    source_df  = pd.read_excel(source_file,  usecols=read_filter)
    updated_df = pd.read_excel(updated_file, usecols=read_filter)

    missing = False
    for label, df in [("source", source_df), ("output", updated_df)]:
        if KEY_COLUMN not in df.columns:
            logger.warning(f"  '{KEY_COLUMN}' 列が {label} ファイルに見つかりません: {source_file.name}")
            missing = True
            break
    if missing:
        continue

    # 管理番号をキーに内部結合
    merged = source_df.merge(
        updated_df,
        on=KEY_COLUMN,
        suffixes=("_before", "_after"),
        how="inner",
    )

    # コンテキスト列は _after を優先して統合
    for col in CONTEXT_COLUMNS:
        after_col  = f"{col}_after"
        before_col = f"{col}_before"
        if after_col in merged.columns:
            merged[col] = merged[after_col]
        elif before_col in merged.columns:
            merged[col] = merged[before_col]

    # 変更行を検出
    changed_mask = pd.Series(False, index=merged.index)
    for col in TARGET_COLUMNS:
        changed_mask |= merged.apply(
            lambda r, c=col: values_differ(r.get(f"{c}_before"), r.get(f"{c}_after")),
            axis=1,
        )

    diff_merged = merged[changed_mask]

    # 管理番号ごとに変更カラムセットを構築（機能1で使用）
    changed_lookup: dict[str, set] = {}
    for _, row in diff_merged.iterrows():
        key = row[KEY_COLUMN]
        changed_cols = set()
        for col in TARGET_COLUMNS:
            if values_differ(row.get(f"{col}_before"), row.get(f"{col}_after")):
                changed_cols.add(col)
        changed_lookup[key] = changed_cols

    if not changed_lookup:
        logger.info(f"  変更なし ({source_file.name})")
        continue

    logger.info(f"  変更行: {len(changed_lookup):,} 行")
    red_font = Font(color="FFFF0000")

    # ----------------------------------------------------------------
    # 機能1: output/<ファイル名>.xlsx の変更セルを直接赤字にする
    # ----------------------------------------------------------------
    logger.info(f"  [機能1] output/{updated_file.name} の変更セルを赤字にします")

    output_wb = load_workbook(updated_file)
    output_ws = output_wb.active
    header_idx = {cell.value: cell.column for cell in output_ws[1]}
    key_col_idx = header_idx.get(KEY_COLUMN)

    if key_col_idx is None:
        logger.warning(f"  '{KEY_COLUMN}' 列が workbook に見つかりません")
    else:
        highlighted = 0
        for row in output_ws.iter_rows(min_row=2, max_row=output_ws.max_row):
            key_val = row[key_col_idx - 1].value
            if key_val not in changed_lookup:
                continue
            for col in changed_lookup[key_val]:
                col_idx = header_idx.get(col)
                if col_idx:
                    row[col_idx - 1].font = red_font
                    highlighted += 1
        output_wb.save(updated_file)
        logger.info(f"  ✓ {highlighted} セルを赤字にしました -> output/{updated_file.name}")

    output_wb.close()

    # ----------------------------------------------------------------
    # 機能2: _変更箇所.xlsx を作成する
    # ----------------------------------------------------------------
    logger.info(f"  [機能2] {diff_output_file.name} を作成します")

    out_cols  = [KEY_COLUMN]
    out_cols += [c for c in CONTEXT_COLUMNS if c in diff_merged.columns]
    for col in TARGET_COLUMNS:
        out_cols += [f"{col}_before", f"{col}_after"]

    rename_map = {}
    for col in TARGET_COLUMNS:
        rename_map[f"{col}_before"] = f"変更前_{col}"
        rename_map[f"{col}_after"]  = f"変更後_{col}"

    diff_df = diff_merged[out_cols].rename(columns=rename_map).copy()

    if diff_output_file.exists():
        diff_output_file.unlink()

    diff_df.to_excel(diff_output_file, index=False, sheet_name="差分")

    # _変更箇所.xlsx の変更後セルも赤字に
    diff_wb = load_workbook(diff_output_file)
    diff_ws = diff_wb.active
    diff_header = {cell.value: idx for idx, cell in enumerate(diff_ws[1], 1)}
    diff_red = 0
    for row_idx in range(2, diff_ws.max_row + 1):
        for col in TARGET_COLUMNS:
            b_col = f"変更前_{col}"
            a_col = f"変更後_{col}"
            if b_col not in diff_header or a_col not in diff_header:
                continue
            before_val = diff_ws.cell(row=row_idx, column=diff_header[b_col]).value
            after_cell = diff_ws.cell(row=row_idx, column=diff_header[a_col])
            if values_differ(before_val, after_cell.value):
                after_cell.font = red_font
                diff_red += 1
    diff_wb.save(diff_output_file)
    diff_wb.close()
    logger.info(f"  ✓ {len(diff_df):,} 行、{diff_red} セルを赤字にしました -> output/{diff_output_file.name}")

logger.info("=" * 80)
logger.info("差分処理完了")
logger.info("=" * 80)
