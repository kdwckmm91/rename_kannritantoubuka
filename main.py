#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台帳の管理担当部課を一括変換するメインスクリプト

機能:
    - 対応関係ファイル（mapping_master.xlsx）から対応データを読み込む
    - input/ 配下の台帳ファイルを一括処理して output/ に同名出力
    - 台帳名と同名シートがある場合は（設置場所, 設置場所詳細）でも突合
    - source/ に同名の神様データがある場合のみ差分を赤字化して出力
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil
import logging
from datetime import datetime

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ファイルパス設定
BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / 'input'
SOURCE_DIR = BASE_DIR / 'source'
OUTPUT_DIR = BASE_DIR / 'output'
MAPPING_FILE = BASE_DIR / 'mapping_master.xlsx'

# 特殊ルール: (組織管理区分１名, 利用者名) → (部課コード, 部課名)
SPECIAL_RULES: dict = {
    ('衛星', '増田　昌史'): (822108000004, '事業企画部'),
}

# 上書きルール: (対象列名, 値) → (部課コード, 部課名)
OVERRIDE_RULES: list = [
    ('組織管理区分１名', 'RAN品',  822108000171, '無線ネットワークデザイン部門'),
    ('組織管理区分１名', '無特',   822108020008, 'RANプロダクト推進部門'),
    ('組織管理区分２名', '4GvRAN', 822108020004, 'RANインテグレーション部門'),
]
OVERRIDE_DICT: dict = {
    (col, val): (code, name)
    for col, val, code, name in OVERRIDE_RULES
}

# 差分処理定数
KEY_COLUMN      = '管理番号'
CONTEXT_COLUMNS = ['設置場所', '組織管理区分１名']
TARGET_COLUMNS  = ['利用者名', '管理担当部課', '管理担当部課名']
READ_COLUMNS    = [KEY_COLUMN] + CONTEXT_COLUMNS + TARGET_COLUMNS
LOCATION_SHEET_REQUIRED_COLUMNS = [
    '設置場所',
    '設置場所詳細',
    '新しい管理担当部課コード',
    '新しい管理担当部課名',
]

# アーカイブ対象（旧版・調査用・一回限りのスクリプト）
ARCHIVE_FILES = [
    '_rewrite_replace_user_names.py',
    'check_satellite.py',
    'check_satellite_detail.py',
    'check_structure.py',
    'create_diff_highlight.py',
    'create_macro_template.py',
    'create_mapping.py',
    'extract_names.py',
    'generate_mapping.py',
    'restore_user_names.py',
    'update_all_ledgers.py',
    'update_ledgers_org_priority.py',
    'update_by_org_category.py',
]


def organize_files():
    """使わない Python ファイルを archive/ フォルダに移動"""
    archive_dir = BASE_DIR / 'archive'
    archive_dir.mkdir(exist_ok=True)
    moved = []
    for filename in ARCHIVE_FILES:
        src = BASE_DIR / filename
        if src.exists():
            dst = archive_dir / filename
            shutil.move(str(src), str(dst))
            moved.append(filename)
    if moved:
        logger.info(f"アーカイブ移動：{len(moved)} 件 → archive/")
        for f in moved:
            logger.info(f"  移動: {f}")
    else:
        logger.info("アーカイブ対象ファイルなし")


def load_mapping():
    """対応関係シートと台帳名対応の設置場所シートを読み込む"""
    logger.info(f"対応関係ファイルを読み込み中: {MAPPING_FILE}")
    
    if not MAPPING_FILE.exists():
        logger.error(f"対応関係ファイルが見つかりません: {MAPPING_FILE}")
        raise FileNotFoundError(f"対応関係ファイルが見つかりません: {MAPPING_FILE}")
    
    try:
        mapping_df = pd.read_excel(MAPPING_FILE, sheet_name='対応関係')
        logger.info(f"✓ 対応関係シートを読み込みました: {len(mapping_df)} 行")

        location_sheet_map = {}
        with pd.ExcelFile(MAPPING_FILE) as excel_file:
            for sheet_name in excel_file.sheet_names:
                if sheet_name == '対応関係':
                    continue
                sheet_df = pd.read_excel(excel_file, sheet_name=sheet_name)
                location_dict = build_location_mapping(sheet_df, sheet_name)
                if location_dict:
                    location_sheet_map[sheet_name] = location_dict
                    logger.info(
                        f"✓ 設置場所突合シートを読み込みました: {sheet_name} ({len(location_dict)} 件)"
                    )

        logger.info(f"✓ 設置場所突合シート総数: {len(location_sheet_map)}")
        return mapping_df, location_sheet_map
    except Exception as e:
        logger.error(f"対応関係ファイルの読み込みエラー: {e}")
        raise

def build_location_mapping(sheet_df: pd.DataFrame, sheet_name: str) -> dict:
    """設置場所シートから (設置場所, 設置場所詳細) の対応辞書を作成"""
    missing_cols = [
        col for col in LOCATION_SHEET_REQUIRED_COLUMNS
        if col not in sheet_df.columns
    ]
    if missing_cols:
        logger.warning(
            f"設置場所突合シートをスキップします: {sheet_name} (不足列: {missing_cols})"
        )
        return {}

    location_dict = {}
    for _, row in sheet_df.iterrows():
        location = str(row['設置場所']).strip() if pd.notna(row.get('設置場所')) else ''
        detail = str(row['設置場所詳細']).strip() if pd.notna(row.get('設置場所詳細')) else ''
        new_code = row['新しい管理担当部課コード']
        new_name = row['新しい管理担当部課名']

        # 設置場所は空白許容（設置場所詳細のみで突合したい行を扱う）
        if not detail or pd.isna(new_code) or pd.isna(new_name):
            continue

        try:
            new_code = int(new_code)
        except (ValueError, TypeError):
            logger.warning(
                f"部課コードが数値に変換できません ({sheet_name}): "
                f"設置場所={location}, 設置場所詳細={detail}, 値={new_code}"
            )
            continue

        location_dict[(location, detail)] = (new_code, str(new_name).strip())

    return location_dict


def get_input_ledger_files() -> list:
    """input/ 配下の処理対象 .xlsx ファイルを取得（Excel一時ファイルは除外）"""
    if not INPUT_DIR.exists():
        logger.warning(f'入力フォルダが見つかりません: {INPUT_DIR}')
        return []

    ledger_files = [
        file
        for file in sorted(INPUT_DIR.glob('*.xlsx'))
        if not file.name.startswith('~$')
    ]
    logger.info(f'入力台帳ファイル検出: {len(ledger_files)} 件')
    return ledger_files

def update_ledger(ledger_df, mapping_df, location_dict=None):
    """台帳を対応関係ファイルで更新"""
    logger.info("台帳の更新を開始...")
    location_dict = location_dict or {}

    # 突合キー種別ごとの辞書を構築（除外フラグが立っている行はスキップ）
    user_dict = {}   # 利用者名 → (code, name)
    org1_dict = {}   # 組織管理区分１名 → (code, name)
    org2_dict = {}   # 組織管理区分２名 → (code, name)
    excluded_count = 0

    for _, row in mapping_df.iterrows():
        # 除外フラグが設定されている行はスキップ
        if pd.notna(row.get('除外フラグ')) and str(row['除外フラグ']).strip() not in ('', 'nan'):
            excluded_count += 1
            continue

        key      = str(row['キー']).strip()
        new_code = row['新しい管理担当部課コード']
        new_name = row['新しい管理担当部課名']
        key_type = str(row.get('突合キー種別', '利用者名')).strip()

        if pd.notna(new_code):
            try:
                new_code = int(new_code)
            except (ValueError, TypeError):
                logger.warning(f"部課コードが数値に変換できません: {new_code}")
                continue

        if key_type == '組織管理区分１名':
            org1_dict[key] = (new_code, new_name)
        elif key_type == '組織管理区分２名':
            org2_dict[key] = (new_code, new_name)
        else:
            user_dict[key] = (new_code, new_name)

    logger.info(f"  利用者名マッピング:          {len(user_dict)} 件")
    logger.info(f"  組織管理区分１名マッピング:  {len(org1_dict)} 件")
    logger.info(f"  組織管理区分２名マッピング:  {len(org2_dict)} 件")
    logger.info(f"  除外（フラグあり）:          {excluded_count} 件")

    # 更新対象行をトラッキング
    cnt_special  = 0
    cnt_override = 0
    cnt_location = 0
    cnt_org1     = 0
    cnt_org2     = 0
    cnt_user     = 0
    unmatched_count = 0

    has_org2_col = '組織管理区分２名' in ledger_df.columns
    has_location_cols = (
        '設置場所' in ledger_df.columns and '設置場所詳細' in ledger_df.columns
    )

    # 各行について処理
    for idx, row in ledger_df.iterrows():
        user_name = str(row['利用者名']).strip()         if pd.notna(row.get('利用者名'))         else None
        org1_name = str(row['組織管理区分１名']).strip() if pd.notna(row.get('組織管理区分１名')) else None
        org2_name = str(row['組織管理区分２名']).strip() if has_org2_col and pd.notna(row.get('組織管理区分２名')) else None
        location = str(row['設置場所']).strip() if has_location_cols and pd.notna(row.get('設置場所')) else ''
        detail = str(row['設置場所詳細']).strip() if has_location_cols and pd.notna(row.get('設置場所詳細')) else ''

        matched = False

        # 優先1: SPECIAL_RULES（組織管理区分１名 × 利用者名の組合せ）
        if not matched and org1_name and user_name:
            result = SPECIAL_RULES.get((org1_name, user_name))
            if result:
                ledger_df.at[idx, '管理担当部課']  = result[0]
                ledger_df.at[idx, '管理担当部課名'] = result[1]
                cnt_special += 1
                matched = True

        # 優先2: OVERRIDE_RULES（組織管理区分１名 → ２名 の順）
        if not matched:
            for col_name, val in [('組織管理区分１名', org1_name), ('組織管理区分２名', org2_name)]:
                if val:
                    result = OVERRIDE_DICT.get((col_name, val))
                    if result:
                        ledger_df.at[idx, '管理担当部課']  = result[0]
                        ledger_df.at[idx, '管理担当部課名'] = result[1]
                        cnt_override += 1
                        matched = True
                        break

        # 優先3: mapping_master 組織管理区分１名
        if not matched and org1_name and org1_name in org1_dict:
            new_code, new_name = org1_dict[org1_name]
            ledger_df.at[idx, '管理担当部課']  = new_code
            ledger_df.at[idx, '管理担当部課名'] = new_name
            cnt_org1 += 1
            matched = True

        # 優先4: mapping_master 組織管理区分２名
        if not matched and org2_name and org2_name in org2_dict:
            new_code, new_name = org2_dict[org2_name]
            ledger_df.at[idx, '管理担当部課']  = new_code
            ledger_df.at[idx, '管理担当部課名'] = new_name
            cnt_org2 += 1
            matched = True

        # 優先5: mapping_master 利用者名（除外フラグ付きはマッピング辞書から除外済み）
        if not matched and user_name and user_name in user_dict:
            new_code, new_name = user_dict[user_name]
            ledger_df.at[idx, '管理担当部課']  = new_code
            ledger_df.at[idx, '管理担当部課名'] = new_name
            cnt_user += 1
            matched = True

        # 優先6: 台帳同名シートの設置場所突合（設置場所 × 設置場所詳細）
        if not matched and location_dict and detail:
            result = location_dict.get((location, detail))
            if result:
                ledger_df.at[idx, '管理担当部課'] = result[0]
                ledger_df.at[idx, '管理担当部課名'] = result[1]
                cnt_location += 1
                matched = True

        if not matched:
            unmatched_count += 1

    total = cnt_special + cnt_override + cnt_location + cnt_org1 + cnt_org2 + cnt_user
    logger.info(f'✓ 台帳の更新完了: {total:,} 行')
    logger.info(f'  - 特殊ルール（SPECIAL）:         {cnt_special}')
    logger.info(f'  - ハードコード上書き（OVERRIDE）: {cnt_override}')
    logger.info(f'  - 設置場所突合（同名シート）:      {cnt_location}')
    logger.info(f'  - 組織管理区分１名突合:           {cnt_org1}')
    logger.info(f'  - 組織管理区分２名突合:           {cnt_org2}')
    logger.info(f'  - 利用者名突合:                   {cnt_user}')
    logger.info(f'  - 非突合行数（変更なし）:         {unmatched_count}')

    return ledger_df

def save_ledger(ledger_df, output_file):
    """更新した台帳を新しいExcelファイルとして保存"""
    logger.info(f"更新した台帳を保存中: {output_file}")
    
    try:
        # 元のExcelファイルと同じフォーマットを保持するため、
        # openpyxl を使用して行・列のフォーマットを保持
        ledger_df.to_excel(output_file, index=False, sheet_name='Sheet1', engine='openpyxl')
        logger.info(f"✓ ファイルを保存しました: {output_file}")
        return True
    except Exception as e:
        logger.error(f"ファイル保存エラー: {e}")
        raise


def values_differ(left, right) -> bool:
    """NaN を考慮した値の比較"""
    if pd.isna(left) and pd.isna(right):
        return False
    if pd.isna(left) or pd.isna(right):
        return True
    return left != right


def create_diff(source_file: Path, updated_file: Path) -> None:
    """
    source と output を管理番号キーで比較し差分処理を行う。
    [機能1] output/<名前>.xlsx の変更セルを直接赤字にする
    [機能2] output/<名前>_変更箇所.xlsx に変更行のみ抽出して出力する
    """
    logger.info(f'  差分処理: {source_file.name}')

    read_filter = lambda col: col in READ_COLUMNS
    source_df  = pd.read_excel(source_file,  usecols=read_filter)
    updated_df = pd.read_excel(updated_file, usecols=read_filter)

    for label, df in [('source', source_df), ('output', updated_df)]:
        if KEY_COLUMN not in df.columns:
            logger.warning(f"  '{KEY_COLUMN}' 列が {label} ファイルに見つかりません")
            return

    merged = source_df.merge(
        updated_df,
        on=KEY_COLUMN,
        suffixes=('_before', '_after'),
        how='inner',
    )

    for col in CONTEXT_COLUMNS:
        after_col  = f'{col}_after'
        before_col = f'{col}_before'
        if after_col in merged.columns:
            merged[col] = merged[after_col]
        elif before_col in merged.columns:
            merged[col] = merged[before_col]

    changed_mask = pd.Series(False, index=merged.index)
    for col in TARGET_COLUMNS:
        changed_mask |= merged.apply(
            lambda r, c=col: values_differ(r.get(f'{c}_before'), r.get(f'{c}_after')),
            axis=1,
        )

    diff_merged = merged[changed_mask]

    changed_lookup: dict = {}
    for _, row in diff_merged.iterrows():
        key = row[KEY_COLUMN]
        changed_cols = set()
        for col in TARGET_COLUMNS:
            if values_differ(row.get(f'{col}_before'), row.get(f'{col}_after')):
                changed_cols.add(col)
        changed_lookup[key] = changed_cols

    if not changed_lookup:
        logger.info('  変更なし')
        return

    logger.info(f'  変更行: {len(changed_lookup):,} 行')
    red_font = Font(color='FFFF0000')

    # 機能1: output/<名前>.xlsx の変更セルを直接赤字にする
    output_wb  = load_workbook(updated_file)
    output_ws  = output_wb.active
    header_idx = {cell.value: cell.column for cell in output_ws[1]}
    key_col_idx = header_idx.get(KEY_COLUMN)
    if key_col_idx is not None:
        highlighted = 0
        for row in output_ws.iter_rows(min_row=2, max_row=output_ws.max_row):
            key_val = row[key_col_idx - 1].value
            if key_val not in changed_lookup:
                continue
            for col in changed_lookup[key_val]:
                col_idx = header_idx.get(col)
                if col_idx:
                    row[col_idx - 1].font = red_font
                    highlighted += 1
        output_wb.save(updated_file)
        logger.info(f'  [機能1] {highlighted} セルを赤字にしました -> {updated_file.name}')
    output_wb.close()

    # 機能2: _変更箇所.xlsx を作成する
    out_cols  = [KEY_COLUMN]
    out_cols += [c for c in CONTEXT_COLUMNS if c in diff_merged.columns]
    for col in TARGET_COLUMNS:
        out_cols += [f'{col}_before', f'{col}_after']

    rename_map = {}
    for col in TARGET_COLUMNS:
        rename_map[f'{col}_before'] = f'変更前_{col}'
        rename_map[f'{col}_after']  = f'変更後_{col}'

    diff_df     = diff_merged[out_cols].rename(columns=rename_map).copy()
    diff_output = OUTPUT_DIR / source_file.name.replace('.xlsx', '_変更箇所.xlsx')
    if diff_output.exists():
        diff_output.unlink()

    diff_df.to_excel(diff_output, index=False, sheet_name='差分')

    diff_wb  = load_workbook(diff_output)
    diff_ws  = diff_wb.active
    diff_hdr = {cell.value: idx for idx, cell in enumerate(diff_ws[1], 1)}
    diff_red = 0
    for row_idx in range(2, diff_ws.max_row + 1):
        for col in TARGET_COLUMNS:
            b_col = f'変更前_{col}'
            a_col = f'変更後_{col}'
            if b_col not in diff_hdr or a_col not in diff_hdr:
                continue
            before_val = diff_ws.cell(row=row_idx, column=diff_hdr[b_col]).value
            after_cell = diff_ws.cell(row=row_idx, column=diff_hdr[a_col])
            if values_differ(before_val, after_cell.value):
                after_cell.font = red_font
                diff_red += 1
    diff_wb.save(diff_output)
    diff_wb.close()
    logger.info(f'  [機能2] {len(diff_df):,} 行、{diff_red} セルを赤字にしました -> {diff_output.name}')


def main():
    """メイン処理"""
    logger.info("=" * 80)
    logger.info("台帳更新スクリプト開始")
    logger.info(f"実行時刻: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info("=" * 80)
    
    try:
        # 0. 不要ファイルを archive/ に整理
        organize_files()

        # 1. 出力フォルダを作成
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        # 2. 入力台帳を列挙
        ledger_files = get_input_ledger_files()
        if not ledger_files:
            logger.warning('input/ に処理対象の .xlsx ファイルがありません')
            return 1

        # 3. 対応関係ファイルを読み込む
        mapping_df, location_sheet_map = load_mapping()
        logger.info(f'台帳処理対象: {len(ledger_files)} 件')

        diff_executed = 0
        diff_skipped = 0

        # 4. 各台帳ファイルを処理
        for ledger_file in ledger_files:
            logger.info('-' * 60)
            logger.info(f'台帳ファイルを処理: {ledger_file.name}')
            ledger_df = pd.read_excel(ledger_file, sheet_name=0)
            logger.info(f'✓ {len(ledger_df):,} 行を読み込みました')

            sheet_key = ledger_file.stem
            location_dict = location_sheet_map.get(sheet_key, {})
            if location_dict:
                logger.info(
                    f'✓ 同名シート設置場所突合を適用: {sheet_key} ({len(location_dict)} 件)'
                )
            else:
                logger.info(f'同名シート設置場所突合なし: {sheet_key}')

            ledger_df = update_ledger(ledger_df, mapping_df, location_dict=location_dict)
            output_file = OUTPUT_DIR / ledger_file.name
            save_ledger(ledger_df, output_file)

            # 5. source に同名ファイルがある場合のみ差分ファイルを作成
            source_file = SOURCE_DIR / ledger_file.name
            if source_file.exists():
                create_diff(source_file, output_file)
                diff_executed += 1
            else:
                logger.info(
                    f'source/ に同名ファイルがないため差分をスキップ: {ledger_file.name}'
                )
                diff_skipped += 1

        logger.info('-' * 60)
        logger.info(f'差分実行: {diff_executed} 件 / 差分スキップ: {diff_skipped} 件')

        logger.info('=' * 80)
        logger.info('✓ スクリプト完了')
        logger.info('=' * 80)
        
    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())
