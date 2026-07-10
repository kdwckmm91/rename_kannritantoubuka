"""
replace_user_names.py の replacement_rules から replacement_rules.xlsx を生成するスクリプト
一度だけ実行すれば OK。以後は Excel ファイルを直接編集する。
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(r"C:\Temp\ツール\sisantantou\replacement_rules.xlsx")

# replacement_rules をそのまま再現
rows = [
    # 無デザ部門
    ("阿部　友希",   "無デザ",        "後藤　健太"),
    ("小林　満",     "無デザ",        "後藤　健太"),
    ("中田　涼太",   "無デザ",        "後藤　健太"),
    # RANプロダクト部門
    ("谷村　誠",     "RANプロダクト", "川上　純平"),
    ("西岡　裕輝",   "RANプロダクト", "川上　純平"),
    # RANインテグ部門
    ("小田島　祥太", "RANインテグ",   "塚本　裕太"),
    ("加藤　遼介",   "RANインテグ",   "塚本　裕太"),
    ("平井　尊教",   "RANインテグ",   "塚本　裕太"),
    ("内山　忠",     "RANインテグ",   "塚本　裕太"),
    ("近松　慎伍",   "RANインテグ",   "塚本　裕太"),
    ("藤野　伸児",   "RANインテグ",   "塚本　裕太"),
    ("三原　寛高",   "RANインテグ",   "塚本　裕太"),
    ("増田　昌史",   "RANインテグ",   "塚本　裕太"),
    ("井原　泰介",   "RANインテグ",   "塚本　裕太"),
    ("田中　晋也",   "RANインテグ",   "塚本　裕太"),
    ("相良　光毅",   "RANインテグ",   "塚本　裕太"),
    ("相原　聖",     "RANインテグ",   "塚本　裕太"),
    ("吉村　匠二",   "RANインテグ",   "塚本　裕太"),
    ("高瀬　健太",   "RANインテグ",   "塚本　裕太"),
    ("平本　義貴",   "RANインテグ",   "塚本　裕太"),
    ("竹内　悠",     "RANインテグ",   "塚本　裕太"),
    ("橋本　謙一",   "RANインテグ",   "塚本　裕太"),
    ("斎藤　健太郎", "RANインテグ",   "塚本　裕太"),
    ("高原　将紀",   "RANインテグ",   "塚本　裕太"),
    ("中村　徹",     "RANインテグ",   "塚本　裕太"),
    ("松谷　龍一",   "RANインテグ",   "塚本　裕太"),
    ("相野　宏文",   "RANインテグ",   "塚本　裕太"),
    ("坂井　信行",   "RANインテグ",   "塚本　裕太"),
    ("大場　弘和",   "RANインテグ",   "塚本　裕太"),
    ("小西　顕也",   "RANインテグ",   "塚本　裕太"),
    ("大野　禎久",   "RANインテグ",   "塚本　裕太"),
    ("古沢　祐之",   "RANインテグ",   "塚本　裕太"),
    ("鬼頭　理香",   "RANインテグ",   "塚本　裕太"),
    ("宮崎　寛之",   "RANインテグ",   "塚本　裕太"),
    ("山崎　肇",     "RANインテグ",   "塚本　裕太"),
    ("山田　太郎",   "RANインテグ",   "塚本　裕太"),
    ("岡本　誠一郎", "RANインテグ",   "塚本　裕太"),
    ("安藤　英浩",   "RANインテグ",   "塚本　裕太"),
    ("長嶋　嶺",     "RANインテグ",   "塚本　裕太"),
    ("森岡　康史",   "RANインテグ",   "塚本　裕太"),
    ("北澤　伸一郎", "RANインテグ",   "塚本　裕太"),
    ("宮野　勇人",   "RANインテグ",   "塚本　裕太"),
    ("住岡　直美",   "RANインテグ",   "塚本　裕太"),
    ("彦坂　諭志",   "RANインテグ",   "塚本　裕太"),
    ("須貝　宜嗣",   "RANインテグ",   "塚本　裕太"),
    ("田中　文子",   "RANインテグ",   "塚本　裕太"),
    ("山野　滉之",   "RANインテグ",   "塚本　裕太"),
    ("島野　大雅",   "RANインテグ",   "塚本　裕太"),
    ("磯辺　浩子",   "RANインテグ",   "塚本　裕太"),
]

df = pd.DataFrame(rows, columns=["置換前利用者名", "部門", "置換後利用者名"])
df["除外フラグ"] = None   # 空欄 = 置換対象、1 = 対象外（既にいない / 異動予定）

df.to_excel(OUTPUT, index=False, sheet_name="置換ルール", engine="openpyxl")
# ※ 列ヘッダー「除外フラグ」はそのまま維持（スクリプトとの整合性を保つ）

# --- 見た目を整える ---
wb = load_workbook(OUTPUT)
ws = wb.active

# ヘッダー書式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 列幅を自動調整
col_widths = {"A": 18, "B": 16, "C": 16, "D": 30}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# D1 の列ヘッダーはそのまま「除外フラグ」を維持（上書きしない）

# 交互に薄いグレー背景
gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
for row_idx in range(2, ws.max_row + 1):
    if row_idx % 2 == 0:
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).fill = gray_fill

# 枠線
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
    for cell in row:
        cell.border = border

wb.save(OUTPUT)
print(f"✓ {OUTPUT.name} を生成しました（{len(df)} 件）")
print(f"  列: {df.columns.tolist()}")
print(f"  除外フラグ: 空欄 = 置換対象、1 = 対象外（既にいない / 異動予定）")
