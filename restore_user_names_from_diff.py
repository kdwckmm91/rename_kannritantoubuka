#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""変更箇所ファイルをもとに利用者名を復元するスクリプト。

処理内容:
- output/無デザ_無シ技_台帳_変更箇所.xlsx から
    変更後_利用者名 が対象利用者名に一致する行を抽出
- 管理番号をキーに、output/無デザ_無シ技_台帳.xlsx の利用者名を
  変更前_利用者名に戻す
- 元ファイルは上書きせず、新規ファイルとして出力する
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl import load_workbook


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


BASE_DIR = Path(__file__).parent
OUTPUT_DIR = BASE_DIR / "output"
DIFF_FILE = OUTPUT_DIR / "無デザ_無シ技_台帳_変更箇所.xlsx"
LEDGER_FILE = OUTPUT_DIR / "無デザ_無シ技_台帳.xlsx"
RESTORED_FILE = OUTPUT_DIR / "無デザ_無シ技_台帳_利用者名復元.xlsx"

TARGET_AFTER_USERS = ["小田島　祥太", "谷村　誠"]

DIFF_REQUIRED_COLUMNS = ["管理番号", "変更前_利用者名", "変更後_利用者名"]
LEDGER_REQUIRED_COLUMNS = ["管理番号", "利用者名"]


def normalize_key(value) -> str:
    """管理番号比較用にキーを正規化する。"""
    if pd.isna(value):
        return ""
    return str(value).strip()


def ensure_file_exists(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {path}")


def ensure_columns(df: pd.DataFrame, required_columns: list[str], source_name: str) -> None:
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"{source_name} に必須列が不足しています: {missing}")


def build_restore_map(diff_df: pd.DataFrame) -> Dict[str, str]:
    """変更箇所ファイルから 管理番号 -> 変更前_利用者名 の復元マップを構築する。"""
    target_users = {name.strip() for name in TARGET_AFTER_USERS}
    filtered = diff_df.loc[
        diff_df["変更後_利用者名"].astype(str).str.strip().isin(target_users),
        ["管理番号", "変更前_利用者名"],
    ].copy()

    if filtered.empty:
        logger.warning("対象行が見つかりません: 変更後_利用者名 in %s", TARGET_AFTER_USERS)
        return {}

    filtered["_key"] = filtered["管理番号"].map(normalize_key)
    filtered["_before"] = filtered["変更前_利用者名"].fillna("").astype(str).str.strip()
    filtered = filtered[filtered["_key"] != ""]

    # 同一管理番号で変更前利用者名が複数ある場合は先頭値を採用して警告
    duplicated = filtered[filtered.duplicated("_key", keep=False)]
    if not duplicated.empty:
        for key, key_df in duplicated.groupby("_key"):
            unique_before = [v for v in key_df["_before"].unique() if v != ""]
            if len(unique_before) > 1:
                logger.warning(
                    "同一管理番号で変更前_利用者名が複数あります: 管理番号=%s, 値=%s",
                    key,
                    unique_before,
                )

    restore_df = filtered.drop_duplicates("_key", keep="first")
    restore_map = dict(zip(restore_df["_key"], restore_df["_before"]))

    logger.info("復元対象キー数: %d", len(restore_map))
    return restore_map


def restore_user_names(restore_map: Dict[str, str]) -> tuple[int, int, int]:
    """台帳ファイルに復元値を反映し、新規ファイルとして保存する。"""
    wb = load_workbook(LEDGER_FILE)
    ws = wb.active

    header = {cell.value: cell.column for cell in ws[1]}
    for col in LEDGER_REQUIRED_COLUMNS:
        if col not in header:
            wb.close()
            raise ValueError(f"台帳ファイルに必須列が不足しています: {col}")

    key_col = header["管理番号"]
    user_col = header["利用者名"]

    updated = 0
    already_same = 0
    matched_keys = set()

    for row_idx in range(2, ws.max_row + 1):
        key_raw = ws.cell(row=row_idx, column=key_col).value
        key = normalize_key(key_raw)
        if key == "" or key not in restore_map:
            continue

        matched_keys.add(key)
        current_user = ws.cell(row=row_idx, column=user_col).value
        current_user_str = "" if current_user is None else str(current_user).strip()
        restore_user = restore_map[key]

        if current_user_str == restore_user:
            already_same += 1
            continue

        ws.cell(row=row_idx, column=user_col).value = restore_user
        updated += 1

    unmatched = len(set(restore_map.keys()) - matched_keys)

    wb.save(RESTORED_FILE)
    wb.close()

    return updated, already_same, unmatched


def main() -> int:
    logger.info("=" * 80)
    logger.info("利用者名復元スクリプト開始")
    logger.info("対象変更後_利用者名: %s", ", ".join(TARGET_AFTER_USERS))
    logger.info("=" * 80)

    try:
        ensure_file_exists(DIFF_FILE)
        ensure_file_exists(LEDGER_FILE)

        diff_df = pd.read_excel(DIFF_FILE, sheet_name=0)
        ensure_columns(diff_df, DIFF_REQUIRED_COLUMNS, "変更箇所ファイル")

        restore_map = build_restore_map(diff_df)
        if not restore_map:
            logger.warning("復元対象がないため処理を終了します")
            return 0

        updated, already_same, unmatched = restore_user_names(restore_map)

        logger.info("復元完了")
        logger.info("  更新件数: %d", updated)
        logger.info("  既に同値: %d", already_same)
        logger.info("  管理番号未一致: %d", unmatched)
        logger.info("  出力ファイル: %s", RESTORED_FILE)
        logger.info("=" * 80)
        return 0

    except Exception as exc:
        logger.error("エラーが発生しました: %s", exc)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
