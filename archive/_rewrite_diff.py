"""create_diff_optimized.py を管理番号キー結合版に書き換えるスクリプト"""
from pathlib import Path

NEW_CONTENT = '''\
"""
差分ファイル生成スクリプト（管理番号キー結合版）

source/ と output/ で同名の .xlsx ファイルを自動検出して比較。
「管理番号」列をキーに内部結合し、TARGET_COLUMNS が異なる行のみを抽出。
変更後の値を赤字にした差分ファイルを output/<元ファイル名>_変更箇所.xlsx として出力する。

事前に update_ledgers.py を実行して output/ にファイルを生成しておくこと。
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BASE_DIR   = Path(r"C:\\Temp\\ツール\\sisantantou")
SOURCE_DIR = BASE_DIR / "source"
OUTPUT_DIR = BASE_DIR / "output"

KEY_COLUMN      = "管理番号"
CONTEXT_COLUMNS = ["設置場所", "利用者名", "組織管理区分１名"]
TARGET_COLUMNS  = ["管理担当部課", "管理担当部課名"]
READ_COLUMNS    = [KEY_COLUMN] + CONTEXT_COLUMNS + TARGET_COLUMNS


def values_differ(left, right) -> bool:
    if pd.isna(left) and pd.isna(right):
        return False
    if pd.isna(left) or pd.isna(right):
        return True
    return left != right


logger.info("=" * 80)
logger.info("差分ファイルの作成開始（管理番号キー結合版）")
logger.info("=" * 80)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# source/ の .xlsx のうち output/ に同名ファイルがあるペアを自動検出
pairs = []
for src in sorted(SOURCE_DIR.glob("*.xlsx")):
    dst = OUTPUT_DIR / src.name
    if dst.exists():
        diff_out = OUTPUT_DIR / src.name.replace(".xlsx", "_変更箇所.xlsx")
        pairs.append((src, dst, diff_out))

if not pairs:
    logger.warning("比較対象のファイルペアが見つかりません。")
    logger.warning("先に update_ledgers.py を実行して output/ にファイルを生成してください。")
else:
    logger.info(f"{len(pairs)} 件のファイルペアを検出")

for source_file, updated_file, output_file in pairs:
    logger.info(f"\\n処理中: {source_file.name}")

    read_filter = lambda col: col in READ_COLUMNS
    source_df  = pd.read_excel(source_file,  usecols=read_filter)
    updated_df = pd.read_excel(updated_file, usecols=read_filter)

    missing = False
    for label, df in [("source", source_df), ("output", updated_df)]:
        if KEY_COLUMN not in df.columns:
            logger.warning(f"  ⚠ '{KEY_COLUMN}' 列が {label} ファイルに見つかりません: {source_file.name}")
            missing = True
            break
    if missing:
        continue

    # 管理番号をキーに内部結合
    merged = source_df.merge(
        updated_df,
        on=KEY_COLUMN,
        suffixes=("_before", "_after"),
        how="inner",
    )

    # コンテキスト列は _after を優先して統合
    for col in CONTEXT_COLUMNS:
        after_col  = f"{col}_after"
        before_col = f"{col}_before"
        if after_col in merged.columns:
            merged[col] = merged[after_col]
        elif before_col in merged.columns:
            merged[col] = merged[before_col]

    # 変更行を検出
    changed_mask = pd.Series(False, index=merged.index)
    for col in TARGET_COLUMNS:
        changed_mask |= merged.apply(
            lambda r, c=col: values_differ(r.get(f"{c}_before"), r.get(f"{c}_after")),
            axis=1,
        )

    diff_df = merged[changed_mask].copy()

    if diff_df.empty:
        logger.info(f"  変更なし ({source_file.name})")
        continue

    # 出力列を組み立て
    out_cols  = [KEY_COLUMN]
    out_cols += [c for c in CONTEXT_COLUMNS if c in diff_df.columns]
    for col in TARGET_COLUMNS:
        out_cols += [f"{col}_before", f"{col}_after"]

    rename_map = {}
    for col in TARGET_COLUMNS:
        rename_map[f"{col}_before"] = f"変更前_{col}"
        rename_map[f"{col}_after"]  = f"変更後_{col}"

    diff_df = diff_df[out_cols].rename(columns=rename_map)

    if output_file.exists():
        output_file.unlink()

    diff_df.to_excel(output_file, index=False, sheet_name="差分")
    logger.info(f"  ✓ ファイルを出力: {output_file.name}")

    # 変更後の値を赤字に
    output_wb = load_workbook(output_file)
    output_ws = output_wb.active
    red_font  = Font(color="FFFF0000")

    header_idx = {cell.value: idx for idx, cell in enumerate(output_ws[1], 1)}
    changed_cells = 0
    for row_idx in range(2, output_ws.max_row + 1):
        for col in TARGET_COLUMNS:
            before_col = f"変更前_{col}"
            after_col  = f"変更後_{col}"
            if before_col not in header_idx or after_col not in header_idx:
                continue
            before_val = output_ws.cell(row=row_idx, column=header_idx[before_col]).value
            after_cell = output_ws.cell(row=row_idx, column=header_idx[after_col])
            if values_differ(before_val, after_cell.value):
                after_cell.font = red_font
                changed_cells += 1

    output_wb.save(output_file)
    output_wb.close()
    logger.info(f"  ✓ {len(diff_df):,} 行、{changed_cells} セルを赤字にしました")

logger.info(f"\\n{'='*80}")
logger.info("✓ 差分ファイルの作成完了")
logger.info(f"{'='*80}")
'''

target = Path(r"C:\Temp\ツール\sisantantou\create_diff_optimized.py")
target.write_text(NEW_CONTENT, encoding="utf-8")
print(f"✓ {target.name} を書き換えました ({len(NEW_CONTENT)} chars)")
