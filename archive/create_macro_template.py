#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel マクロ用テンプレートファイル（.xlsm）を作成
"""

import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

def create_macro_template():
    """
    マクロ対応の Excel テンプレートを作成
    
    ※ 注意: openpyxl は VBA コードを埋め込めないため、
    このファイルには構造のみ作成され、VBA コードは手動で埋め込む必要があります。
    """
    
    template_file = Path(r"C:\Temp\ツール\sisantantou\マクロ用テンプレート.xlsm")
    
    # 新しいワークブックを作成
    wb = openpyxl.Workbook()
    
    # デフォルトシートをメインシートに名前変更
    ws_main = wb.active
    ws_main.title = "メイン"
    
    # マッピングシートを追加
    ws_mapping = wb.create_sheet("マッピング")
    
    # マッピングシートにヘッダーを設定
    ws_mapping['A1'].value = 'キー（利用者名または組織管理区分１名）'
    ws_mapping['B1'].value = '新しい管理担当部課コード'
    ws_mapping['C1'].value = '新しい管理担当部課名'
    
    # ひな形データを設定
    ws_mapping['A2'].value = '利用者A'
    ws_mapping['B2'].value = 822108020001
    ws_mapping['C2'].value = '営業部'
    
    ws_mapping['A3'].value = '利用者B'
    ws_mapping['B3'].value = 822108020002
    ws_mapping['C3'].value = '企画部'
    
    ws_mapping['A4'].value = '組織管理区分X'
    ws_mapping['B4'].value = 822108020003
    ws_mapping['C4'].value = '技術部'
    
    # 列幅を設定
    ws_mapping.column_dimensions['A'].width = 35
    ws_mapping.column_dimensions['B'].width = 20
    ws_mapping.column_dimensions['C'].width = 20
    
    # メインシートに説明を追加
    ws_main['A1'].value = '注意: このファイルにマクロを埋め込む必要があります'
    ws_main['A2'].value = 'ステップ:'
    ws_main['A3'].value = '1. Alt + F11 でマクロエディタを開く'
    ws_main['A4'].value = '2. Project > Insert > Module で新しいモジュールを追加'
    ws_main['A5'].value = '3. macro_template.vba の内容をコピーして貼り付け'
    ws_main['A6'].value = '4. ファイルを保存（Ctrl + S）'
    ws_main['A7'].value = '5. マクロを実行'
    
    ws_main.column_dimensions['A'].width = 60
    
    # ファイルを保存
    try:
        # .xlsx として保存
        xlsx_file = Path(r"C:\Temp\ツール\sisantantou\マクロ用テンプレート.xlsx")
        wb.save(xlsx_file)
        print(f"✓ テンプレートファイルを作成しました: {xlsx_file}")
        print(f"\n注意: このファイルは .xlsx 形式です。")
        print(f"VBA マクロを使用するには以下の手順に従ってください:")
        print(f"1. このファイルを開く")
        print(f"2. 『マッピング』シートに対応関係データを入力")
        print(f"3. macro_template.vba のコードをマクロエディタに貼り付け")
        print(f"4. ファイルを .xlsm 形式で保存")
        print(f"5. マクロを実行")
        
    except Exception as e:
        print(f"エラー: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    create_macro_template()
