import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

BASE_DIR = Path(r"C:\Temp\ツール\sisantantou")
SOURCE_DIR = BASE_DIR / "source"

file_pairs = [
    (
        SOURCE_DIR / "無デザ_無シ技_台帳_20260618.xlsx",
        BASE_DIR / "無デザ_無シ技_台帳_20260618_updated.xlsx",
        BASE_DIR / "無デザ_無シ技_台帳_変更箇所.xlsx"
    ),
    (
        SOURCE_DIR / "RAN技_無シ技_台帳_20260618.xlsx",
        BASE_DIR / "RAN技_無シ技_台帳_20260618_updated.xlsx",
        BASE_DIR / "RAN技_無シ技_台帳_変更箇所.xlsx"
    )
]

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
red_font = None

logger.info("=" * 80)
logger.info("差分ファイルの作成開始")
logger.info("=" * 80)

for source_file, updated_file, output_file in file_pairs:
    if not source_file.exists() or not updated_file.exists():
        logger.warning(f"ファイルが見つかりません: {source_file.name} または {updated_file.name}")
        continue

    logger.info(f"\n処理中: {source_file.name}")

    # 元ファイルと更新ファイルを読み込み
    source_df = pd.read_excel(source_file, sheet_name=0)
    updated_df = pd.read_excel(updated_file, sheet_name=0)

    # 更新ファイルをコピーして出力ファイルを作成
    updated_df.to_excel(output_file, index=False, sheet_name='Sheet1')

    # openpyxlで変更箇所に色を付ける
    wb = load_workbook(output_file)
    ws = wb.active

    change_count = 0

    # データ行をチェック（ヘッダーはrow 1）
    for row_idx in range(2, len(updated_df) + 2):
        for col_idx, col_name in enumerate(updated_df.columns, 1):
            # 管理担当部課と管理担当部課名列のみをチェック
            if col_name not in ['管理担当部課', '管理担当部課名']:
                continue

            source_val = source_df.iloc[row_idx - 2][col_name]
            updated_val = updated_df.iloc[row_idx - 2][col_name]

            # 値が異なる場合、セルを赤色で塗りつぶし
            if pd.isna(source_val) != pd.isna(updated_val) or (pd.notna(source_val) and pd.notna(updated_val) and source_val != updated_val):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = red_fill
                change_count += 1

    wb.save(output_file)
    logger.info(f"✓ {change_count}セルが変更されました")
    logger.info(f"✓ 出力ファイル: {output_file.name}")

logger.info(f"\n{'='*80}")
logger.info("✓ 差分ファイルの作成完了")
logger.info(f"{'='*80}")
